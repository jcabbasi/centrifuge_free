import numpy as np
from numpy.core.numeric import NaN
import pandas as pd
import os
from openpyxl import load_workbook


class DataStorage:
    
    def __init__(self):
        self.DF=pd.DataFrame(index=[], columns=['Omega','Time','P','Sw','Pc','RF','IterNo'])
       
    def assign(self,omega,t,P,Sw,Pc,RF,IterNo):
        xx=len(self.DF)
        
        self.DF.loc[xx+1,'Omega']=omega
        self.DF.loc[xx+1,'Time']=t
        self.DF.loc[xx+1,'P']=P
        self.DF.loc[xx+1,'Sw']=Sw
        self.DF.loc[xx+1,'Pc']=Pc
        self.DF.loc[xx+1,'RF']=RF
        self.DF.loc[xx+1,'IterNo']=IterNo
    
    
    
    def ExtractSimulationResults(self,ColList):
        
        exportvector=np.zeros([  len(self.DF.loc[:,'Time']) , len(ColList) ])
        jj=0
        for ii in ColList:
            exportvector[:,jj]=self.DF.loc[:,ii]
            jj=jj+1
        
        return exportvector
    
    def SavetoPickle(self,filename):
        fn= "./"+filename+".pkl"
        self.DF.to_pickle(fn)
    
    def SaveRFtoCSV(self,filename):
        fn= "./"+filename+".csv"
        self.DF.loc[:,['Time','RF']].to_csv(fn)

    
  
class EquilDataStorage:
    
    def __init__(self,gn,omegaVect):
        self.DF=pd.DataFrame(index=[], columns=['Omega','Time','StageTime','AnalyticalRF','P','Sw','Pc','RF'])
        omegaVectN=omegaVect.copy(); omegaVectA=omegaVect.copy()
        for ii in range(len(omegaVect)):
            omegaVectN[ii]='Num'+str(omegaVectN[ii]); omegaVectA[ii]='Ana'+str(omegaVectA[ii])           
        self.SW=pd.DataFrame(index=range(gn),columns=omegaVectA+omegaVectN )
        

       
    def assign(self,omega,t,StageTime,AnalyticalRF,P,Sw,Pc,RF,AnalyticalSW):
        xx=len(self.DF)
        
        self.DF.loc[xx+1,'Omega']=omega
        self.DF.loc[xx+1,'Time']=t
        self.DF.loc[xx+1,'P']=P
        self.DF.loc[xx+1,'Sw']=Sw
        self.DF.loc[xx+1,'Pc']=Pc
        self.DF.loc[xx+1,'RF']=RF
        self.DF.loc[xx+1,'StageTime']=StageTime
        self.DF.loc[xx+1,'AnalyticalRF']=AnalyticalRF
        self.SW['Ana'+str(omega)]=AnalyticalSW        
        self.SW['Num'+str(omega)]=Sw
        
        
            

        

    
    
    
    def ExtractSimulationResults(self,ColList):
        
        exportvector=np.zeros([  len(self.DF.loc[:,'Time']) , len(ColList) ])
        jj=0
        for ii in ColList:
            exportvector[:,jj]=self.DF.loc[:,ii]
            jj=jj+1
        
        return exportvector
    
    def SavetoPickle(self,filename):
        fn= "./"+filename+".pkl"
        self.DF.to_pickle(fn)
    
    def SavetoCSV(self,filename):
        fn= "./"+filename+".csv"
        fnSats= "./"+filename+"_SW.csv"
        
        self.DF.loc[:,['Omega','Time','AnalyticalRF','RF',]].to_csv(fn)
        self.SW.to_csv(fnSats)

        
        

        
class LoadandSave:
    
    
    
    def LoadDataFromExcel(filename,fromCell, toCell):
        
        absolutepath = os.path.abspath(__file__)
        fileDirectory = os.path.dirname(absolutepath)

        loc = fileDirectory+filename
        # To open Workbook
        wb = load_workbook(loc)
        ws = wb['data']
        observedX=ws[fromCell:toCell]
        observedData=np.zeros(np.shape(observedX))
        for ii in range(np.shape(observedX)[0]):
            for jj in range(np.shape(observedX)[1]):
                observedData[ii,jj]=observedX[ii][jj].internal_value  

        return observedData
    
        

    
        